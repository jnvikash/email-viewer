import json
import logging
import os
import sqlite3
import threading
from datetime import datetime, timezone
from pathlib import Path

from .db import DB_PATH, get_db
from .msg_parser import parse_msg_metadata, get_msg_body, list_attachments

log = logging.getLogger(__name__)

_lock = threading.Lock()
_threads: dict[int, threading.Thread] = {}  # user_id -> thread

MAX_FTS_BODY = 1_048_576  # 1 MB


def _extract_attachment_text(file_path: str) -> str:
    """Best-effort text extraction from attachments for FTS indexing."""
    texts = []
    try:
        from .msg_parser import list_attachments, get_attachment_bytes
        for att in list_attachments(file_path):
            fname = att["filename"].lower()
            try:
                result = get_attachment_bytes(file_path, att["attach_index"])
                if not result:
                    continue
                _, _, data = result
                if fname.endswith(".pdf"):
                    texts.append(_extract_pdf(data))
                elif fname.endswith(".docx"):
                    texts.append(_extract_docx(data))
                elif fname.endswith((".xlsx", ".xls")):
                    texts.append(_extract_xlsx(data))
                elif fname.endswith((".txt", ".csv", ".html", ".htm")):
                    texts.append(data.decode("utf-8", errors="replace")[:50_000])
            except Exception as e:
                log.debug("Attachment text extract failed %s: %s", fname, e)
    except Exception as e:
        log.debug("list_attachments failed for %s: %s", file_path, e)
    return " ".join(t for t in texts if t)[:MAX_FTS_BODY]


def _extract_pdf(data: bytes) -> str:
    try:
        import io
        from pdfminer.high_level import extract_text
        return extract_text(io.BytesIO(data))[:100_000]
    except Exception:
        return ""


def _extract_docx(data: bytes) -> str:
    try:
        import io
        from docx import Document
        doc = Document(io.BytesIO(data))
        return "\n".join(p.text for p in doc.paragraphs)[:100_000]
    except Exception:
        return ""


def _extract_xlsx(data: bytes) -> str:
    try:
        import io
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        parts = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                parts.append(" ".join(str(c) for c in row if c is not None))
                if len(parts) > 5000:
                    break
        return "\n".join(parts)[:100_000]
    except Exception:
        return ""


def _scan_files(root: str) -> list[str]:
    result = []
    for dirpath, _, filenames in os.walk(root):
        for fn in filenames:
            if fn.lower().endswith(".msg"):
                result.append(os.path.join(dirpath, fn))
    return result


def _folder_path(file_path: str, root: str) -> str:
    rel = os.path.relpath(os.path.dirname(file_path), root)
    return rel.replace("\\", "/")


def _update_state(conn: sqlite3.Connection, user_id: int, **kwargs):
    kwargs["updated_at"] = datetime.now(timezone.utc).isoformat()
    sets = ", ".join(f"{k} = ?" for k in kwargs)
    conn.execute(f"UPDATE index_state SET {sets} WHERE user_id = ?", list(kwargs.values()) + [user_id])
    conn.commit()


class IndexerThread(threading.Thread):
    def __init__(self, user_id: int, root_path: str, app_context):
        super().__init__(daemon=True, name=f"IndexerThread-user{user_id}")
        self.user_id = user_id
        self.root_path = root_path
        self.app_context = app_context

    def run(self):
        with self.app_context:
            conn = sqlite3.connect(str(DB_PATH))
            conn.execute("PRAGMA journal_mode=WAL")
            conn.execute("PRAGMA busy_timeout=5000")
            conn.execute("PRAGMA foreign_keys=ON")
            try:
                self._run(conn)
            except Exception as e:
                log.exception("Indexer crashed for user %d: %s", self.user_id, e)
                _update_state(conn, self.user_id, status="error", error_msg=str(e), finished_at=_now())
            finally:
                conn.close()
                with _lock:
                    if self.user_id in _threads:
                        del _threads[self.user_id]

    def _run(self, conn: sqlite3.Connection):
        _update_state(conn, self.user_id, status="running", started_at=_now(), done_files=0, skipped=0, error_msg=None)

        log.info("Scanning %s for .msg files (user %d)...", self.root_path, self.user_id)
        files = _scan_files(self.root_path)
        total = len(files)
        log.info("Found %d .msg files for user %d", total, self.user_id)
        _update_state(conn, self.user_id, total_files=total)

        done = 0
        skipped = 0
        BATCH = 50

        for i in range(0, total, BATCH):
            batch = files[i:i + BATCH]
            rows_emails = []
            rows_attachments = []
            rows_fts = []

            for fp in batch:
                try:
                    meta = parse_msg_metadata(fp)
                    if meta is None:
                        skipped += 1
                        continue

                    folder = _folder_path(fp, self.root_path)
                    _, text_body = get_msg_body(fp)
                    att_text = _extract_attachment_text(fp)
                    attachments = list_attachments(fp)

                    rows_emails.append((
                        self.user_id, fp, folder,
                        meta["subject"], meta["sender_name"], meta["sender_email"],
                        json.dumps(meta["recipients"]), meta["date_sent"],
                        _now(), meta["has_html_body"], meta["body_preview"],
                        meta["attachment_count"], meta["size_bytes"], meta["importance"],
                    ))
                    rows_attachments.append((fp, attachments))
                    rows_fts.append((fp, meta["subject"], meta["sender_name"], meta["sender_email"],
                                     text_body[:MAX_FTS_BODY], att_text))
                    done += 1
                except Exception as e:
                    log.warning("Skip %s: %s", fp, e)
                    skipped += 1

            # Write batch in one transaction
            try:
                conn.execute("BEGIN IMMEDIATE")
                conn.executemany(
                    """INSERT OR IGNORE INTO emails
                       (user_id, file_path, folder_path, subject, sender_name, sender_email,
                        recipients, date_sent, date_indexed, has_html_body, body_preview,
                        attachment_count, size_bytes, importance)
                       VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    rows_emails,
                )

                for fp, atts in rows_attachments:
                    row = conn.execute("SELECT id FROM emails WHERE file_path=? AND user_id=?", (fp, self.user_id)).fetchone()
                    if row:
                        eid = row[0]
                        existing = conn.execute("SELECT COUNT(*) FROM attachments WHERE email_id=?", (eid,)).fetchone()[0]
                        if existing == 0:
                            conn.executemany(
                                "INSERT INTO attachments (email_id, filename, mime_type, size_bytes, attach_index) VALUES (?,?,?,?,?)",
                                [(eid, a["filename"], a["mime_type"], a["size"], a["attach_index"]) for a in atts],
                            )

                for fp, subj, sname, semail, body, att_text in rows_fts:
                    row = conn.execute("SELECT id FROM emails WHERE file_path=? AND user_id=?", (fp, self.user_id)).fetchone()
                    if row:
                        eid = row[0]
                        existing = conn.execute("SELECT COUNT(*) FROM fts_body WHERE email_id=?", (eid,)).fetchone()[0]
                        if existing == 0:
                            conn.execute(
                                "INSERT INTO fts_body (email_id, body_text, attachment_text) VALUES (?,?,?)",
                                (eid, body, att_text),
                            )
                            conn.execute(
                                "INSERT INTO emails_fts (rowid, subject, sender_name, sender_email, body_text, attachment_text) VALUES (?,?,?,?,?,?)",
                                (eid, subj, sname, semail, body, att_text),
                            )

                _update_state(conn, self.user_id, done_files=done, skipped=skipped)
                conn.commit()
            except Exception as e:
                conn.rollback()
                log.error("Batch write failed for user %d: %s", self.user_id, e)

        _update_state(conn, self.user_id, status="done", done_files=done, skipped=skipped, finished_at=_now())
        log.info("Indexing complete for user %d: %d indexed, %d skipped", self.user_id, done, skipped)


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def start_indexing(user_id: int, root_path: str, app) -> str:
    """Start background indexing for a user."""
    with _lock:
        if user_id in _threads and _threads[user_id].is_alive():
            return "already_running"
        _threads[user_id] = IndexerThread(user_id, root_path, app.app_context())
        _threads[user_id].start()
        return "started"


def get_status(user_id: int) -> dict:
    """Get indexing status for a user."""
    try:
        conn = sqlite3.connect(str(DB_PATH))
        conn.row_factory = sqlite3.Row
        row = conn.execute("SELECT * FROM index_state WHERE user_id=?", (user_id,)).fetchone()
        conn.close()
        if row:
            return dict(row)
    except Exception:
        pass
    return {"status": "idle", "total_files": 0, "done_files": 0, "skipped": 0}


def reset_index(user_id: int) -> None:
    """Reset index for a user."""
    with _lock:
        if user_id in _threads and _threads[user_id].is_alive():
            return
        conn = sqlite3.connect(str(DB_PATH))
        # Delete user's emails (cascades to attachments/fts)
        conn.execute("DELETE FROM emails WHERE user_id=?", (user_id,))
        conn.execute("DELETE FROM fts_body WHERE email_id NOT IN (SELECT id FROM emails)")
        conn.execute("DELETE FROM emails_fts WHERE rowid NOT IN (SELECT id FROM emails)")
        conn.execute("UPDATE index_state SET status='idle', total_files=0, done_files=0, skipped=0, started_at=NULL, finished_at=NULL, error_msg=NULL WHERE user_id=?", (user_id,))
        conn.commit()
        conn.close()

